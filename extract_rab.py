import openpyxl
import os
import sys
from io import BytesIO

def extract_rab(file_input, sheet_name="RAB", start_row=77):
    # file_input can be a string path or a file-like object
    wb = openpyxl.load_workbook(file_input, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        return None
        
    sheet = wb[sheet_name]
    
    # Try to find "PEKERJAAN" in column 2, rows 1-75
    pekerjaan_name = "Untitled Project"
    for r in range(1, 76):
        cell_val = str(sheet.cell(row=r, column=2).value or "").strip().lower()
        if "pekerjaan" in cell_val:
            val = sheet.cell(row=r, column=4).value
            if val:
                pekerjaan_name = str(val).strip().lstrip(':').strip()
                break
    
    # Deteksi Header Otomatis di baris 76
    header_row = 76
    headers = {}
    for cell in sheet[header_row]:
        val = str(cell.value).strip().lower() if cell.value else ""
        if val:
            headers[val] = cell.column
            
    # Target Mapping Aliases
    target_fields = {
        "item": ["item pekerjaan", "uraian pekerjaan", "uraian", "pekerjaan"],
        "satuan": ["satuan", "unit"],
        "volume": ["volume", "vol"],
        "harga": ["harga satuan", "harga", "unit price"]
    }
    
    col_map = {}
    for field, aliases in target_fields.items():
        for alias in aliases:
            if alias in headers:
                col_map[field] = headers[alias]
                break
    
    # Fallback jika tidak ketemu (berdasarkan screenshot bos)
    # Screenshot menunjukkan: Col 3=Item, Col 4=Satuan, Col 5=Volume, Col 6=Harga
    if "item" not in col_map: col_map["item"] = 3
    if "satuan" not in col_map: col_map["satuan"] = 4
    if "volume" not in col_map: col_map["volume"] = 5
    if "harga" not in col_map: col_map["harga"] = 6
    
    data = []
    for row_idx in range(start_row, sheet.max_row + 1):
        if sheet.row_dimensions[row_idx].hidden:
            continue
            
        item = sheet.cell(row=row_idx, column=col_map["item"]).value
        satuan = sheet.cell(row=row_idx, column=col_map["satuan"]).value
        volume = sheet.cell(row=row_idx, column=col_map["volume"]).value
        harga = sheet.cell(row=row_idx, column=col_map["harga"]).value
        
        if item:
            item_str = str(item).strip()
            if item_str.lower() == "jumlah":
                continue
                
            item_str = item_str.replace("m²", "m2").replace("m³", "m3")
            satuan_str = (str(satuan).strip() if satuan else "").replace("m²", "m2").replace("m³", "m3")
                
            data.append({
                "item": item_str,
                "satuan": satuan_str,
                "volume": volume,
                "harga": harga,
                "pajak": 11,
                "keterangan": "",
                "kunci": "FALSE" if harga is not None and harga != "" else "TRUE"
            })
            
    final_data = []
    if pekerjaan_name:
        final_data.append({
            "item": pekerjaan_name,
            "satuan": "", "volume": None, "harga": None, "pajak": "", "keterangan": "", "kunci": ""
        })
    final_data.extend(data)
    return final_data

def create_xlsx_buffer(data):
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active
    new_sheet.title = "Extracted RAB"
    headers = ["Item Pekerjaan", "Satuan", "Volume", "Harga Satuan", "Pajak", "Keterangan", "Kunci"]
    new_sheet.append(headers)
    for row in data:
        new_sheet.append([row["item"], row["satuan"], row["volume"], row["harga"], row["pajak"], row["keterangan"], row["kunci"]])
    buffer = BytesIO()
    new_wb.save(buffer)
    buffer.seek(0)
    return buffer
